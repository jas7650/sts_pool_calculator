import math
import os
import sys
from bs4 import BeautifulSoup
import openpyxl
from tracker.utils.sheet_utils import *
from tracker.utils.scrape_utils import *
from objects.Team import Team
from objects.Player import Player
from objects.Pool import Pool
import subprocess
import argparse


def main():
    parser = argparse.ArgumentParser(
        description="Utility for calculating pools given an html document",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument(
        '-t',
        dest="tournament",
        default=None,
        required=True,
        help="Input to specify what tournament to generate pools for [default: %(default)s]"
    )
    parser.add_argument(
        '-m',
        dest="major",
        default=False,
        required=False,
        action="store_true",
        help="Input to specify that the tournament is a major [default: %(default)s]"
    )

    args = parser.parse_args(args=None if sys.argv[1:] else ['-h'])

    file = args.tournament
    is_major = args.major

    cmd = 'python ./tracker/update_sheet.py'
    p = subprocess.Popen(cmd)
    p.wait()

    html_doc = readFile(file)
    soup = BeautifulSoup(html_doc, 'html.parser')
    teams = getTeams(soup)
    teams = sorted(teams, key=lambda x: x.getPoints(), reverse=True)
    
    createPoolsWorkbook(teams, is_major, file)


def createPoolsWorkbook(teams, is_major, file):
    wb = getWorkBook(str(os.path.basename(file)).split(".")[0])
    wb = removeSheets(wb)
    createTeamsSheet(teams, wb)
    createPoolsSheets(teams, is_major, wb)


def createTeamsSheet(teams, wb):
    teamNames = ['Team']
    playerOnes = ['Player One']
    playerTwos = ['Player Two']
    playerOnePoints = ['Player One Points']
    playerTwoPoints = ['Player Two Points']
    teamPoints = ['Team Points']

    for team in teams:
        teamNames.append(team.getTeamName())
        playerOnes.append(team.getPlayerOne().getName())
        playerOnePoints.append(team.getPlayerOne().getPoints())
        playerTwos.append(team.getPlayerTwo().getName())
        playerTwoPoints.append(team.getPlayerTwo().getPoints())
        teamPoints.append(team.getPlayerOne().getPoints() + team.getPlayerTwo().getPoints())
    data = [teamNames, playerOnes, playerOnePoints, playerTwos, playerTwoPoints, teamPoints]
    wb = writeToSheet(data, wb, "Teams")

    saveWorkBook(wb, 'pool_calculations.xlsx')


def createPoolsSheets(teams, is_major, wb):
    num_power_pools = 0
    if is_major:
        num_power_pools = getNumPowerPools(teams)
        min_pools = math.ceil((len(teams)-(num_power_pools*4))/5)
        max_pools = math.floor((len(teams)-(num_power_pools*4))/4)
    else:
        min_pools = math.ceil(len(teams)/8)
        max_pools = math.floor(len(teams)/4)
    
    num_pools = min_pools
    
    teams, power_pools = getPowerPools(teams, num_power_pools)
    
    while num_pools <= max_pools:
        pools = getPools(teams, num_pools, num_power_pools)
        wb = writePoolsToSheet(wb, power_pools, pools)
        num_pools += 1
    saveWorkBook(wb, 'pool_calculations.xlsx')


def getPoolNum(num_pools : int, index : int, offset : int):
    iteration = int(math.floor(index/num_pools))
    if iteration % 2 == 0:
        pool_num = (index%num_pools)
    else:
        pool_num = num_pools-(index%num_pools)-1
    return pool_num + offset + 1


def getPowerPools(teams : list, num_power_pools : int):
    power_pools = {}
    for i in range(num_power_pools):
        power_pools[i+1] = Pool(i+1)
    for i in range(num_power_pools*4):
        pool_num = getPoolNum(num_power_pools, i, 0)
        team = teams.pop(0)
        power_pools[pool_num].addTeam(team)
    return teams, power_pools


def getPools(teams : list, num_pools : int, num_power_pools : int):
    pools = {}
    teams_copy = teams.copy()
    team_num = 0
    for i in range(num_pools):
        pools[i+1+num_power_pools] = Pool(i+1+num_power_pools)
    while len(teams_copy) > 0:
        pool_num = getPoolNum(num_pools, team_num, num_power_pools)
        # print(f"Pool Num: {pool_num}")
        team = teams_copy.pop(0)
        pools[pool_num].addTeam(team)
        team_num += 1
    return pools


def writePoolsToSheet(wb, power_pools, pools):
    pool_nums = ['Pool']
    teamNames = ['Team Name']
    playerOnes = ['Player One']
    playerTwos = ['Player Two']
    playerOnePoints = ['Player One Points']
    playerTwoPoints = ['Player Two Points']
    teamPoints = ['Team Points']
    for pool in power_pools.keys():
        for team in power_pools[pool].getTeams():
            pool_nums.append(power_pools[pool].getPoolNum())
            teamNames.append(team.getTeamName())
            playerOnes.append(team.getPlayerOneName())
            playerTwos.append(team.getPlayerTwoName())
            playerOnePoints.append(team.getPlayerOne().getPoints())
            playerTwoPoints.append(team.getPlayerTwo().getPoints())
            teamPoints.append(team.getPoints())
        pool_nums.append("")
        teamNames.append("")
        playerOnes.append("")
        playerTwos.append("")
        playerOnePoints.append("")
        playerTwoPoints.append("")
        teamPoints.append("")

    for pool in pools.keys():
        for team in pools[pool].getTeams():
            pool_nums.append(pools[pool].getPoolNum())
            teamNames.append(team.getTeamName())
            playerOnes.append(team.getPlayerOneName())
            playerTwos.append(team.getPlayerTwoName())
            playerOnePoints.append(team.getPlayerOne().getPoints())
            playerTwoPoints.append(team.getPlayerTwo().getPoints())
            teamPoints.append(team.getPoints())
        pool_nums.append("")
        teamNames.append("")
        playerOnes.append("")
        playerTwos.append("")
        playerOnePoints.append("")
        playerTwoPoints.append("")
        teamPoints.append("")
    data = [pool_nums, teamNames, playerOnes, playerTwos, playerOnePoints, playerTwoPoints, teamPoints]
    wb = writeToSheet(data, wb, f"{len(power_pools) + len(pools)} Pools")
    return wb


def getTeams(soup):
    teams = []
    teamNames = []
    playerOnes = []
    playerTwos = []
    points = getPoints()

    for team in soup.findAll('div', attrs = {'class':'team-name'}):
        teamNames.append(team.text)

    for team in soup.findAll('div', attrs = {'class':'players'}):
        players = team.text
        andIndex = players.find(" and ")
        playerOnes.append(Player(cleanName(players[0:andIndex]), getPlayerPoints(cleanName(players[0:andIndex]), points)))
        playerTwos.append(Player(cleanName(players[andIndex+5:]), getPlayerPoints(cleanName(players[andIndex+5:]), points)))

    for i in range(len(teamNames)):
        team = Team(teamNames[i], playerOnes[i], playerTwos[i])

        teams.append(team)
    return teams


def getNumPowerPools(teams):
    if len(teams) < 9:
        return 0
    if len(teams) > 8 and len(teams) < 24:
        return 1
    if len(teams) > 23 and len(teams) < 44:
        return 2
    if len(teams) > 43 and len(teams) < 64:
        return 3
    if len(teams) > 63:
        return 4


def getPlayerPoints(player, points):
    if player in points.keys():
        return points[player]
    return 0

 
def getPoints():
    wb = getWorkBook('./tracker/roundnet_season_tracker.xlsx')
    sheet = getSheetByName(wb, "Players")
    points = {}
 
    for i in range(sheet.max_row):
        key = sheet.cell(row=i+1, column=1).value
        points[str(key)] = sheet.cell(row=i+1, column=2).value

    return points


def readFile(file : str):
    with open(file, "r", encoding="utf8") as f:
        html_doc = f.readlines()
    return html_doc[0]

 
if __name__=="__main__":
    main()

