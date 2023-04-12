import openpyxl
import os
import sys


def getWorkBook(filename):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
    return wb


def writeToSheet(data, wb, sheetName):
    if (sheetExists(wb, sheetName)): 
        sheet = getSheetByName(wb, sheetName)
        for col in range(len(data)):
            for row in range(len(data[0])):
                if (row == 0):
                    continue
                writeToCell(row+17, col+1, sheet, data[col][row])
    else:
        sheet = createSheet(wb, sheetName)
        for col in range(len(data)):
            for row in range(len(data[0])):
                writeToCell(row+1, col+1, sheet, data[col][row])
    return wb


def writeToColumn(data, wb, sheetName, col):
    if (sheetExists(wb, sheetName)): 
        sheet = getSheetByName(wb, sheetName)
    else:
        sheet = createSheet(wb, sheetName)
    for row in range(len(data)):
        writeToCell(row+1, col+1, sheet, data[row])
    return wb


def saveWorkBook(wb, filename):
    wb.save(filename)


def getColumnData(sheet, column):
    data = []
    for row in range(2,sheet.max_row+1):
        data.append(getCellValue(row, column, sheet))
    return data


def getRowData(sheet, row):
    data = []
    for column in range(2,sheet.max_column+1):
        data.append(getCellValue(row, column, sheet))
    return data


def getSheetByName(wb, sheetname):
    if sheetExists(wb, sheetname):
        return wb[sheetname]
    print(f'{sheetname} does not exist')
    sys.exit(0)


def sheetExists(wb, sheetname):
    for name in wb.sheetnames:
        if name == sheetname:
            return True
    return False


def removeSheets(wb):
    sheets = wb.sheetnames
    for name in sheets:
        if (name != "Point Distribution 2022"):
            sheet = wb[name]
            wb.remove(sheet)
    return wb


def writeToCell(row, col, sheet, value):
    sheet.cell(row=row, column=col).value = value
    return sheet


def getCellValue(row, col, sheet):
    return sheet.cell(row=row, column=col).value


def createSheet(wb, sheetName):
    return wb.create_sheet(sheetName)
